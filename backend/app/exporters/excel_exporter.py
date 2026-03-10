"""
Excel Exporter - Gera arquivo Excel (.xlsx) formatado das transações
"""

from typing import List, Optional
from app.models.transaction import Transaction


class ExcelExporter:
    """Exporta transações para Excel com formatação profissional."""

    def export(self, transactions: List[Transaction], output_path: str, bank_name: str = "Extrato"):
        """Gera arquivo Excel formatado."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, Fill, PatternFill, Alignment, Border, Side,
                numbers as num_styles
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._export_csv_fallback(transactions, output_path)
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extrato"
        
        # Cores
        COLOR_HEADER = "1E3A5F"
        COLOR_DEBIT = "FFE8E8"
        COLOR_CREDIT = "E8F5E9"
        COLOR_ALT_ROW = "F5F7FA"
        
        # Estilos
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # === TÍTULO ===
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"Extrato Bancário — {bank_name}"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35
        
        # === CABEÇALHO ===
        headers = ["Data", "Descrição", "Tipo", "Débito (R$)", "Crédito (R$)", "Saldo (R$)", "Categoria"]
        col_widths = [14, 45, 10, 16, 16, 16, 18]
        
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[2].height = 25
        
        # === DADOS ===
        for row_idx, t in enumerate(transactions, 3):
            is_odd = (row_idx % 2 == 0)
            row_fill_color = COLOR_DEBIT if t.tipo == "debit" else (COLOR_CREDIT if t.tipo == "credit" else (COLOR_ALT_ROW if is_odd else "FFFFFF"))
            row_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")
            
            data_str = t.data.strftime("%d/%m/%Y") if t.data else ""
            debito = t.valor_abs if t.tipo == "debit" else None
            credito = t.valor_abs if t.tipo == "credit" else None
            
            row_data = [
                data_str,
                t.descricao,
                "Débito" if t.tipo == "debit" else "Crédito",
                debito,
                credito,
                t.saldo,
                t.categoria or ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                
                if col_idx == 1:  # Data
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in (4, 5, 6):  # Valores
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 3:  # Tipo
                    cell.alignment = Alignment(horizontal="center")
                    if t.tipo == "debit":
                        cell.font = Font(name="Calibri", size=10, color="C0392B", bold=True)
                    else:
                        cell.font = Font(name="Calibri", size=10, color="1E8449", bold=True)
        
        # === RESUMO ===
        summary_row = len(transactions) + 4
        ws.cell(row=summary_row, column=1, value="RESUMO").font = Font(bold=True, size=11, color=COLOR_HEADER)
        
        total_debito = sum(t.valor_abs for t in transactions if t.tipo == "debit")
        total_credito = sum(t.valor_abs for t in transactions if t.tipo == "credit")
        saldo = total_credito - total_debito
        
        summary_data = [
            ("Total Transações:", len(transactions)),
            ("Total Débitos:", total_debito),
            ("Total Créditos:", total_credito),
            ("Saldo do Período:", saldo),
        ]
        
        for i, (label, value) in enumerate(summary_data, summary_row + 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
            val_cell = ws.cell(row=i, column=2, value=value)
            val_cell.font = Font(size=10)
            if isinstance(value, float):
                val_cell.number_format = 'R$ #,##0.00'
                if value < 0:
                    val_cell.font = Font(size=10, color="C0392B")
        
        # Congela cabeçalho
        ws.freeze_panes = "A3"
        
        wb.save(output_path)

    def _export_csv_fallback(self, transactions: List[Transaction], output_path: str):
        """Fallback se openpyxl não estiver disponível."""
        import csv
        output_path = output_path.replace(".xlsx", ".csv")
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["Data", "Descrição", "Tipo", "Débito", "Crédito", "Saldo", "Categoria"])
            for t in transactions:
                writer.writerow([
                    t.data.strftime("%d/%m/%Y") if t.data else "",
                    t.descricao,
                    t.tipo,
                    t.valor_abs if t.tipo == "debit" else "",
                    t.valor_abs if t.tipo == "credit" else "",
                    t.saldo or "",
                    t.categoria or ""
                ])
